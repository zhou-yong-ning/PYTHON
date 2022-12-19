import openpyxl
from copy import deepcopy

# 原文：https://www.cnblogs.com/liuda9495/p/9039732.html



def unmergecells(worksheet):
    # 获取所有 合并单元格的 位置信息
    # 是个可迭代对象，单个对象类型：openpyxl.worksheet.cell_range.CellRange
    # print后就是excel坐标信息
    m_list = worksheet.merged_cells

    l = deepcopy(m_list)  # 深拷贝

    # 拆分合并的单元格 并填充内容
    for m_area in l:

        # 这里的行和列的起始值（索引），和Excel的一样，从1开始，并不是从0开始（注意）
        r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col

        worksheet.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
        print('区域:', m_area, '  坐标:', r1, r2, c1, c2)

        # 获取一个单元格的内容
        first_value = worksheet.cell(r1, c1).value

        # 数据填充
        for r in range(r1, r2 + 1):  # 遍历行
            if c2 - c1 > 0:  # 多个列，遍历列
                for c in range(c1, c2 + 1):
                    worksheet.cell(r, c).value = first_value
            else:  # 一个列
                worksheet.cell(r, c1).value = first_value


path = r"C:\Users\zhou\Desktop\pyceshi\新建文件夹\新建 XLSX 工作表.xlsx"
workbook = openpyxl.load_workbook(path)  # 加载excel
name_list = workbook.sheetnames  # 所有sheet的名字
worksheet1 = workbook[name_list[0]]  # 读取第一个工作表
unmergecells(worksheet1)
workbook.save(r"C:\Users\zhou\Desktop\pyceshi\新建文件夹\新建 XLSX 工作表2.xlsx")

